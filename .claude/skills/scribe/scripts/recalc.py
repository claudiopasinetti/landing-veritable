#!/usr/bin/env python3
"""Recalculate Excel formulas via LibreOffice headless.

Usage:
    python3 recalc.py <file.xlsx> [--output <output.xlsx>]
    python3 recalc.py --help

Arguments:
    file.xlsx       Input Excel file to recalculate
    --output, -o    Output file path (default: overwrites input)
    --check         Only check for formula errors, don't recalculate
    --help, -h      Show this help message

Output (JSON):
    {
        "status": "success" | "error",
        "input": "file.xlsx",
        "output": "file.xlsx",
        "total_formulas": 42,
        "total_errors": 0,
        "errors": [],
        "message": ""
    }

Exit codes:
    0 - Success
    1 - User error (file not found, invalid format)
    2 - System error (LibreOffice not found, crash)
"""

import argparse
import json
import os
import shutil
import subprocess
import sys
import tempfile


def find_soffice():
    """Find LibreOffice binary."""
    paths = [
        shutil.which("soffice"),
        shutil.which("libreoffice"),
        "/usr/bin/soffice",
        "/usr/bin/libreoffice",
        "/usr/local/bin/soffice",
        "/snap/bin/libreoffice",
        "/Applications/LibreOffice.app/Contents/MacOS/soffice",
    ]
    for p in paths:
        if p and os.path.isfile(p):
            return p
    return None


def count_formulas(filepath):
    """Count formulas and errors in an xlsx file using openpyxl."""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(filepath, data_only=True)
        total_formulas = 0
        errors = []

        wb_formulas = load_workbook(filepath, data_only=False)
        for sheet_name in wb_formulas.sheetnames:
            ws_f = wb_formulas[sheet_name]
            ws_v = wb[sheet_name]
            for row in ws_f.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        total_formulas += 1
                        val_cell = ws_v[cell.coordinate]
                        if isinstance(val_cell.value, str) and val_cell.value.startswith("#"):
                            errors.append({
                                "sheet": sheet_name,
                                "cell": cell.coordinate,
                                "formula": cell.value,
                                "error": val_cell.value,
                            })
        wb.close()
        wb_formulas.close()
        return total_formulas, errors
    except ImportError:
        return -1, []


def recalc(input_file, output_file=None):
    """Recalculate formulas using LibreOffice."""
    soffice = find_soffice()
    if not soffice:
        return {
            "status": "error",
            "input": input_file,
            "output": None,
            "total_formulas": -1,
            "total_errors": -1,
            "errors": [],
            "message": "LibreOffice not found. Install it to recalculate formulas.",
        }

    if not os.path.isfile(input_file):
        return {
            "status": "error",
            "input": input_file,
            "output": None,
            "total_formulas": -1,
            "total_errors": -1,
            "errors": [],
            "message": f"File not found: {input_file}",
        }

    with tempfile.TemporaryDirectory() as tmpdir:
        # Copy input to temp dir
        tmp_input = os.path.join(tmpdir, os.path.basename(input_file))
        shutil.copy2(input_file, tmp_input)

        # Recalculate via LibreOffice macro
        macro = (
            'macro:///Standard.Module1.Recalc'
        )
        cmd = [
            soffice,
            "--headless",
            "--calc",
            "--convert-to", "xlsx",
            "--outdir", tmpdir,
            tmp_input,
        ]

        try:
            result = subprocess.run(
                cmd, capture_output=True, text=True, timeout=60
            )
        except subprocess.TimeoutExpired:
            return {
                "status": "error",
                "input": input_file,
                "output": None,
                "total_formulas": -1,
                "total_errors": -1,
                "errors": [],
                "message": "LibreOffice timed out after 60 seconds.",
            }

        if result.returncode != 0:
            return {
                "status": "error",
                "input": input_file,
                "output": None,
                "total_formulas": -1,
                "total_errors": -1,
                "errors": [],
                "message": f"LibreOffice error: {result.stderr.strip()}",
            }

        # Determine output path
        recalced = os.path.join(tmpdir, os.path.basename(input_file))
        if not os.path.isfile(recalced):
            # LibreOffice may change extension
            candidates = [f for f in os.listdir(tmpdir) if f.endswith(".xlsx")]
            if candidates:
                recalced = os.path.join(tmpdir, candidates[0])
            else:
                return {
                    "status": "error",
                    "input": input_file,
                    "output": None,
                    "total_formulas": -1,
                    "total_errors": -1,
                    "errors": [],
                    "message": "LibreOffice did not produce output file.",
                }

        # Copy to final output
        final_output = output_file or input_file
        shutil.copy2(recalced, final_output)

        # Count formulas and errors
        total_formulas, errors = count_formulas(final_output)

        return {
            "status": "success",
            "input": input_file,
            "output": final_output,
            "total_formulas": total_formulas,
            "total_errors": len(errors),
            "errors": errors,
            "message": "",
        }


def check_only(input_file):
    """Check for formula errors without recalculating."""
    if not os.path.isfile(input_file):
        return {
            "status": "error",
            "input": input_file,
            "output": None,
            "total_formulas": -1,
            "total_errors": -1,
            "errors": [],
            "message": f"File not found: {input_file}",
        }

    total_formulas, errors = count_formulas(input_file)
    return {
        "status": "success",
        "input": input_file,
        "output": None,
        "total_formulas": total_formulas,
        "total_errors": len(errors),
        "errors": errors,
        "message": "Check only (no recalculation).",
    }


def main():
    parser = argparse.ArgumentParser(
        description="Recalculate Excel formulas via LibreOffice headless."
    )
    parser.add_argument("file", help="Input Excel file (.xlsx)")
    parser.add_argument("-o", "--output", help="Output file path (default: overwrite input)")
    parser.add_argument("--check", action="store_true", help="Only check for errors, don't recalculate")

    args = parser.parse_args()

    if not args.file.endswith((".xlsx", ".xls")):
        result = {
            "status": "error",
            "input": args.file,
            "output": None,
            "total_formulas": -1,
            "total_errors": -1,
            "errors": [],
            "message": "Input must be an Excel file (.xlsx or .xls).",
        }
        print(json.dumps(result, indent=2))
        sys.exit(1)

    if args.check:
        result = check_only(args.file)
    else:
        result = recalc(args.file, args.output)

    print(json.dumps(result, indent=2))
    sys.exit(0 if result["status"] == "success" else 2)


if __name__ == "__main__":
    main()
